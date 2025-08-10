"""
Standalone TenderAI application using Flask with inline templates.
This app allows uploading tender documents (PDF, DOCX, XLSX), extracts
basic information such as EMD, due date, a short summary, and displays
them on a simple dashboard. It stores data in a JSON file and saves
uploaded files into an 'uploads' directory. HTML and CSS are embedded
directly in the Python file via render_template_string to avoid
external template files.
"""

from flask import Flask, request, redirect, url_for, flash, render_template_string, send_from_directory
import os
import uuid
import subprocess
import zipfile
import re
import json
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "secret-key"

# Directory configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
DATA_FILE = os.path.join(BASE_DIR, "tenders.json")

os.makedirs(UPLOAD_DIR, exist_ok=True)

# Load existing tenders from disk if present
if os.path.exists(DATA_FILE):
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        tenders = json.load(f)
else:
    tenders = {}


def save_tenders() -> None:
    """Persist tender data to disk in JSON format."""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(tenders, f, ensure_ascii=False, indent=2)


def extract_text_from_pdf(path: str) -> str:
    """Extract text from a PDF file using the pdftotext command-line tool.
    Returns an empty string if extraction fails.
    """
    try:
        result = subprocess.run(
            ['pdftotext', '-layout', path, '-'],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True
        )
        return result.stdout.decode('utf-8', errors='ignore')
    except Exception as e:
        print(f"PDF extraction failed for {path}: {e}")
        return ""


def extract_text_from_docx(path: str) -> str:
    """Extract text from a DOCX file by reading its XML content.
    Returns an empty string if extraction fails.
    """
    try:
        with zipfile.ZipFile(path) as docx_zip:
            with docx_zip.open('word/document.xml') as document_xml:
                xml = document_xml.read().decode('utf-8', errors='ignore')
                cleaned = re.sub(r'<[^>]+>', '', xml)
        return cleaned
    except Exception as e:
        print(f"DOCX extraction failed for {path}: {e}")
        return ""


def parse_xlsx(path: str):
    """Parse an XLSX file and return a list of rows from the first sheet.
    Each cell is converted to a string; empty cells become empty strings.
    """
    rows = []
    try:
        wb = load_workbook(filename=path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else '' for cell in row])
    except Exception as e:
        print(f"XLSX parsing failed for {path}: {e}")
    return rows


def simple_summary(text: str) -> str:
    """Generate a simple summary by taking the first three sentences."""
    sentences = re.split(r'(?<=[.!?])\s+', text.strip())
    return ' '.join(sentences[:3])


def extract_fields(text: str) -> dict:
    """Extract basic fields like EMD and due date using regex heuristics.
    If not found, fallback patterns are used. Eligibility is taken from the
    beginning of the text.
    """
    fields = {}
    emd_match = re.search(r'(?:EMD|Earnest Money Deposit).*?(₹|Rs\.?|INR)\s*([\d,\.]+)', text, re.IGNORECASE)
    if emd_match:
        fields['emd'] = f"{emd_match.group(1)} {emd_match.group(2)}"
    else:
        general_match = re.search(r'(₹|Rs\.?|INR)\s*([\d,\.]+)', text)
        fields['emd'] = f"{general_match.group(1)} {general_match.group(2)}" if general_match else ''
    date_match = re.search(r'(?:submission(?: date)?|due date|closing date)[^\d]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.IGNORECASE)
    if date_match:
        fields['due_date'] = date_match.group(1)
    else:
        generic_date = re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', text)
        fields['due_date'] = generic_date.group(0) if generic_date else ''
    # Eligibility excerpt: first 200 characters
    fields['eligibility'] = text[:200] + '...' if len(text) > 200 else text
    return fields


def process_file(file_storage) -> dict:
    """Process an uploaded file and return a tender data dictionary."""
    filename = file_storage.filename
    file_id = str(uuid.uuid4())
    save_path = os.path.join(UPLOAD_DIR, file_id + '_' + filename)
    file_storage.save(save_path)

    ext = filename.lower().split('.')[-1]
    text = ''
    table_rows = []
    if ext == 'pdf':
        text = extract_text_from_pdf(save_path)
    elif ext == 'docx':
        text = extract_text_from_docx(save_path)
    elif ext in ['xlsx', 'xls']:
        table_rows = parse_xlsx(save_path)
    # Fallback: if no text but we have rows, join first few rows for summary
    if not text and table_rows:
        flat = '\n'.join([', '.join(row) for row in table_rows[:5]])
        text = flat
    summary = simple_summary(text) if text else ''
    fields = extract_fields(text) if text else {'emd': '', 'due_date': '', 'eligibility': ''}
    tender = {
        'id': file_id,
        'filename': filename,
        'path': save_path,
        'text': text,
        'table_rows': table_rows,
        'summary': summary,
        'emd': fields.get('emd', ''),
        'due_date': fields.get('due_date', ''),
        'eligibility': fields.get('eligibility', ''),
        'uploaded_at': datetime.now().isoformat()
    }
    tenders[file_id] = tender
    save_tenders()
    return tender


# Inline HTML templates with embedded CSS. The {{ rows_html }} and other variables
# are inserted via render_template_string.
INDEX_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TenderAI Dashboard</title>
    <style>
    body {
        font-family: Arial, sans-serif;
        margin: 0;
        padding: 0;
        background-color: #f7f7f7;
        color: #333;
    }
    header {
        background-color: #1f2937;
        color: #fff;
        padding: 1rem;
        text-align: center;
    }
    header h1 { margin: 0; }
    main {
        max-width: 960px;
        margin: 0 auto;
        padding: 1rem;
    }
    section {
        margin-bottom: 2rem;
        padding: 1rem;
        background-color: #fff;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 1rem;
    }
    th, td {
        padding: 0.5rem;
        border: 1px solid #ddd;
        text-align: left;
        font-size: 0.9rem;
    }
    th { background-color: #f3f4f6; }
    .upload-section input[type="file"] {
        padding: 0.5rem;
        margin-right: 0.5rem;
    }
    .upload-section button {
        padding: 0.5rem 1rem;
        background-color: #2563eb;
        color: white;
        border: none;
        border-radius: 4px;
        cursor: pointer;
    }
    .upload-section button:hover { background-color: #1e40af; }
    .hint { font-size: 0.8rem; color: #6b7280; }
    footer {
        text-align: center;
        padding: 1rem 0;
        background-color: #f3f4f6;
        color: #6b7280;
    }
    </style>
</head>
<body>
    <header>
        <h1>TenderAI Dashboard</h1>
    </header>
    <main>
        <section class="upload-section">
            <h2>Upload Tender Documents</h2>
            <form action="/upload" method="post" enctype="multipart/form-data">
                <input type="file" name="files" multiple required>
                <button type="submit">Upload</button>
            </form>
            <p class="hint">Supported formats: PDF, DOCX, XLSX, XLS</p>
        </section>
        <section class="tender-list">
            <h2>Uploaded Tenders</h2>
            <table>
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Filename</th>
                        <th>EMD</th>
                        <th>Due Date</th>
                        <th>Summary</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody>
                    {{ rows_html | safe }}
                </tbody>
            </table>
            <p class="hint">If no rows appear, there are no tenders uploaded yet.</p>
        </section>
    </main>
    <footer>
        <p>&copy; {{ year }} TenderAI</p>
    </footer>
</body>
</html>
"""


DETAIL_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Tender Details - {{ filename }}</title>
    <style>
    body {
        font-family: Arial, sans-serif;
        margin: 0;
        padding: 0;
        background-color: #f7f7f7;
        color: #333;
    }
    header {
        background-color: #1f2937;
        color: #fff;
        padding: 1rem;
        text-align: center;
    }
    nav a {
        color: #fbbf24;
        text-decoration: none;
    }
    main {
        max-width: 960px;
        margin: 0 auto;
        padding: 1rem;
        background-color: #fff;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 1rem;
    }
    th, td {
        padding: 0.5rem;
        border: 1px solid #ddd;
        text-align: left;
        font-size: 0.9rem;
    }
    th { background-color: #f3f4f6; }
    .text-block {
        white-space: pre-wrap;
        background-color: #f9fafb;
        padding: 1rem;
        border-radius: 4px;
        border: 1px solid #e5e7eb;
        max-height: 400px;
        overflow-y: scroll;
    }
    footer {
        text-align: center;
        padding: 1rem 0;
        background-color: #f3f4f6;
        color: #6b7280;
    }
    </style>
</head>
<body>
    <header>
        <h1>Tender Details</h1>
        <nav>
            <a href="/">Back to Dashboard</a>
        </nav>
    </header>
    <main>
        <h2>{{ filename }}</h2>
        <p><strong>Uploaded At:</strong> {{ uploaded_at }}</p>
        <p><strong>EMD:</strong> {{ emd }}</p>
        <p><strong>Due Date:</strong> {{ due_date }}</p>
        <p><strong>Eligibility (excerpt):</strong> {{ eligibility }}</p>
        <h3>Summary</h3>
        <p>{{ summary }}</p>
        {{ table_html | safe }}
        <h3>Full Text</h3>
        <pre class="text-block">{{ full_text }}</pre>
    </main>
    <footer>
        <p>&copy; {{ year }} TenderAI</p>
    </footer>
</body>
</html>
"""


@app.route('/')
def index():
    """Render the dashboard with a list of uploaded tenders."""
    tender_list = sorted(tenders.values(), key=lambda x: x['uploaded_at'], reverse=True)
    rows_html = ''
    for i, tender in enumerate(tender_list, 1):
        summary = tender['summary'] or 'No summary'
        emd = tender['emd'] or '—'
        due = tender['due_date'] or '—'
        rows_html += (
            f"<tr><td>{i}</td><td>{tender['filename']}</td>"
            f"<td>{emd}</td><td>{due}</td><td>{summary}</td>"
            f"<td><a href='/tender/{tender['id']}'>View</a> | "
            f"<a href='/download/{tender['id']}'>Download</a></td></tr>"
        )
    return render_template_string(INDEX_TEMPLATE, rows_html=rows_html, year=datetime.utcnow().year)


@app.route('/upload', methods=['POST'])
def upload():
    """Handle file uploads from the dashboard form."""
    if 'files' not in request.files:
        flash('No file part in the request')
        return redirect(url_for('index'))
    files = request.files.getlist('files')
    if not files or files[0].filename == '':
        flash('No files selected')
        return redirect(url_for('index'))
    for f in files:
        process_file(f)
    flash(f"Uploaded {len(files)} file(s) successfully!")
    return redirect(url_for('index'))


@app.route('/tender/<tid>')
def tender_detail(tid):
    """Show details for a single tender, including summary and table data."""
    tender = tenders.get(tid)
    if not tender:
        flash('Tender not found')
        return redirect(url_for('index'))
    # Build table HTML for XLSX rows if available
    table_html = ''
    if tender.get('table_rows'):
        table_html += '<h3>Table Data (first 10 rows)</h3><table><tbody>'
        for row in tender['table_rows'][:10]:
            cells = ''.join(f"<td>{c}</td>" for c in row[:10])
            table_html += f"<tr>{cells}</tr>"
        table_html += '</tbody></table>'
    return render_template_string(
        DETAIL_TEMPLATE,
        filename=tender['filename'],
        uploaded_at=tender['uploaded_at'],
        emd=tender['emd'] or '—',
        due_date=tender['due_date'] or '—',
        eligibility=tender['eligibility'] or '—',
        summary=tender['summary'] or 'No summary available for this tender.',
        full_text=tender['text'],
        table_html=table_html,
        year=datetime.utcnow().year
    )


@app.route('/download/<tid>')
def download(tid):
    """Serve the original uploaded file for download."""
    tender = tenders.get(tid)
    if not tender:
        flash('Tender not found')
        return redirect(url_for('index'))
    directory = os.path.dirname(tender['path'])
    filename = os.path.basename(tender['path'])
    return send_from_directory(directory, filename, as_attachment=True)


if __name__ == '__main__':
    # Run the app on all interfaces to accommodate hosting platforms
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))